# encoding=utf8
from openpyxl import load_workbook
from openpyxl import utils as xl_u
import time
from datetime import date, timedelta, datetime
import smtplib
import win32com.client
# import unicodecsv as csv
import csv
import os, os.path
import requests
import sys
import logging
import jaydebeapi as jdbc
import gspread
import pygsheets
import httplib2
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
import oauth2client
from oauth2client import file
from oauth2client import client
from oauth2client import tools
from oauth2client.service_account import ServiceAccountCredentials
import io
import math
from jaydebeapi import _DEFAULT_CONVERTERS
import pandas as pd
import numpy as np
from pytz import timezone
import shutil
import re
# Gdrive
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from pydrive.files import GoogleDriveFileList
import ast
import zipfile
from time import gmtime
# Handle Error
from googleapiclient.errors import HttpError
# SKYPE
from skpy import Skype, SkypeEventLoop, SkypeNewMessageEvent

TARGET_TIMEZONE = 'Asia/Singapore'
LOCAL_TIMEZONE = 'Asia/Jakarta'

### CHECK SUCCESS MARKER FOR THESE DEPENDENCIES BEFORE START RUNNING
DEPENDENCY_LIST = ['regional_db', 'user_profile']

### FOR PRESTO CONNECTION
PRESTO_USER = '' ## erased for confidentiality
PRESTO_PASSWORD = '' ## erased for confidentiality
JAR_PATH = os.getcwd() + '\PN - User List\presto-jdbc-0.205.jar'
PRESTO_SCHEMA = 'shopee_id'
PRESTO_SCHEMA_ANLYS = 'shopee_id_anlys'
RETRY = 5

### QUERY RESULT AS CSV OUTPUT
QUERY_LIST = ['pn_user_list.sql']

# Define scopes and client secret file for downloading sheet
scopes = 'https://www.googleapis.com/auth/drive'
application_name = 'Google drive API Python Quickstart'
client_secret_file = 'KEVIN_CLIENT_SERVICE.json'

# Define date and path
today_date = datetime.now().strftime('%Y%m%d %H.%M')
new_report_path = 'G:/My Drive/AUTOMATED_REPORTS/PN - Comm Results/' + today_date + '/' # path
script_path = 'G:/My Drive/AUTOMATED_REPORTS/PN - Comm/PN - User List/' # path
pn_input_filename = 'PN Input ' + today_date

# Define Gdrive folder ID and list folder
folder_id = '1pOUhHvDEb1OcxQHpnRCyOaNfxN7vfZke'
# list_parent_folder = drive.ListFile({'q': "'{}' in parents and trashed=false".format(folder_id)}).GetList()

### FOR WRITING LOG FILE
log_filename = 'PN Script Log'                      
log_formatter = '%(asctime)s - %(levelname)s - %(message)s'
logging.basicConfig(level=logging.INFO, format=log_formatter)
logger = logging.getLogger(__name__)
handler = logging.FileHandler(os.getcwd() + '\PN - User List\\' + log_filename + '.log')
handler.setLevel(logging.INFO)
handler.setFormatter(logging.Formatter(log_formatter))
logger.addHandler(handler)

logger.info('Report date: ' + today_date)

#########################################################

class DictCursor():
    def __init__(self, conn, batch_size=1000):
        self.obj = conn.cursor()
        self.batch_size = batch_size

    def __iter__(self):
        while True:
            headers = tuple(x[0] for x in self.description)
            results = self.obj.fetchmany(self.batch_size)
            if not results:
                break
            for tup in results:
                yield dict(zip(headers, tup))

    def __getattr__(self, attr):
        return self.obj.__getattribute__(attr)


def _to_datetime(rs, col):
    java_val = rs.getString(col)

    if '/' not in java_val:
        java_val = java_val + ' ' + LOCAL_TIMEZONE
    # print(java_val)
    if not java_val:
        return
    d = datetime.datetime.strptime(str(java_val)[:19], "%Y-%m-%d %H:%M:%S")
    return datetime.datetime.strftime(timezone(str(java_val)[24:]).localize(d).astimezone(timezone(TARGET_TIMEZONE)),
                                      "%Y-%m-%d %H:%M:%S")

_DEFAULT_CONVERTERS.update({'TIMESTAMP': _to_datetime})

def remove_illegal_chars(value):
    import string
    illegal_chars = range(0, 32)
    illegal_chars.pop(10)
    illegal_chars_translation = dict((x, None) for x in illegal_chars)
    illegal_chars_string = [chr(x) for x in illegal_chars]
    illegal_chars_string = ''.join(illegal_chars_string)
    emptytrans = string.maketrans('', '')
    if isinstance(value, unicode):
        writevalue = value.translate(illegal_chars_translation)
    elif isinstance(value, basestring):
        writevalue = value.translate(emptytrans, illegal_chars_string)
    else:
        writevalue = value
    return writevalue

def run_presto(query):
    conn = jdbc.connect(jclassname='com.facebook.presto.jdbc.PrestoDriver',
                        # url='jdbc:presto://presto.idata.shopeemobile.com:8443/hive',
                        url='jdbc:presto://presto-secure.idata.shopeemobile.com:443/hive/' + PRESTO_SCHEMA,
                        driver_args={'user': PRESTO_USER, 'password': PRESTO_PASSWORD, 'SSL': 'true'},
                        # 'SSLTrustStorePath': 'presto.presto.idata.shopeemobile.com.truststore.jks',
                        # 'SSLTrustStorePassword': 'shopee-presto'},
                        jars=[JAR_PATH]
                        )

    curs = DictCursor(conn)

    fail = 0
    while fail >= 0 and fail < RETRY:
        try:
            curs.obj._rs = curs._connection.jconn.createStatement().executeQuery(query)
            fail = -1
        except Exception as e:
            fail = fail + 1
            logger.info(str(e))
    if fail == 4:
        logger.info('RETRY 5 times, still error.')
        exit()

    curs.obj._meta = curs.obj._rs.getMetaData()
    desc = curs.description
    col_name = [x[0] for x in desc]
    rows = curs.fetchall()

    return [col_name, rows]

def get_service():
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIAL_FILE, scope)
    gc = gspread.authorize(credentials)
    logger.info('Connect Google Sheet Successfully...')
    return gc

def delete_temp_folder(FolderPath):
    try:
        shutil.rmtree(FolderPath)
    except OSError, e:  ## if failed, report it back to the user ##
        logger.info ("Error: %s - %s." % (e.filename, e.strerror))

# for download pn input sheet
def get_credentials():
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir, 'auto_download.json')
    store = oauth2client.file.Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(client_secret_file, scopes)
        flow.user_agent = application_name
        credentials = tools.run_flow(flow, store)
        logger.info('Storing credentials to ' + credential_path)
    return credentials
        
def get_credentials_path():
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials') 
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir, 'auto_download.json')
    store = oauth2client.file.Storage(credential_path)
    credentials = store.get()
    
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(client_secret_file, scopes)
        flow.user_agent = application_name
        credentials = tools.run_flow(flow, store)
        print('Storing credentials to ' + credential_path)

    return credential_path

def fill_email_excel(NUMBER_OF_SPLITS, i):
    path = 'G:/My Drive/AUTOMATED_REPORTS/PN - Comm/PN - User List/' # path
    filename = 'PN - Automation Email'
    email_excel_path = path + filename + '.xlsx' # Email Template

    logger.info('Email excel path: ' + email_excel_path)
    wb = load_workbook(email_excel_path, data_only=True)
    sheet = wb['Config'] # Configuration Page
    
    sheet['A' + str(i+7)].value = 'ALL'
    sheet['B' + str(i+7)].value = sql_file_name + '_' + str(i+1) + '.csv'
    sheet['C' + str(i+7)].value = 'A+S'
    sheet['D' + str(i+7)].value = sql_file_name + '_' + str(i+1) + '[]'

    wb.save('G:/My Drive/AUTOMATED_REPORTS/PN - Comm/PN - User List/PN - Automation Email.xlsx')

def zip_file():
    zip_results_path = new_report_path + sql_file_name + '.zip' # path
    logger.info('----------------------------------')
    logger.info('Start zipping..')
    logger.info('Change directory to report path..')
    os.chdir(new_report_path)
    print('Current path: ' + new_report_path)
    time.sleep(2)
    file_list = []

    # Get the list of CSV file in the specific folder and get the filename to be zipped
    for file_folder in os.listdir(new_report_path):
        if file_folder.endswith('.csv'):
            file_list.append(file_folder)

    logger.info('File list: ' + str(file_list))
    with zipfile.ZipFile(zip_results_path, 'w') as new_zip:
        logger.info('Current path: ' + os.getcwd())
        for each_file in file_list:
            new_zip.write(each_file, compress_type=zipfile.ZIP_DEFLATED)
    new_zip.close()

def send_message_skype():    
    logger.info('Start sending...')
    username = '' ## erased for confidentiality
    password = '' ## erased for confidentiality
    sk = Skype(username, password) 
    ch = sk.contacts['live:95d6e67964ccccd7'].chat
    # ch2 = sk.contacts['live:95d6e67964ccccd7'].chat

    # Send message
    msg_content = 'Hi, there\'s an error in PN - Comm automation. Please check it. Thank you. - Automated Message'
    ch.sendMsg(msg_content)
    # ch2.sendMsg(msg_content)
    logger.info('Notification sent successfully')
    
def send_email():
    # prepare to send out email!
    msgRoot = MIMEMultipart()
    strFrom = SENDER + ' ' + PERSONAL_EMAIL
    msgRoot['Subject'] = EMAIL_SUBJECT
    msgRoot['From'] = strFrom

    msgRoot['To'] = ", ".join(RECIPIENT_LIST)
    msgText = MIMEText(
        EMAIL_MESSAGE + '</i>', 'html')
    msgRoot.attach(msgText)

    smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login(LOGIN_EMAIL, LOGIN_PASSWORD)
    smtpObj.sendmail(strFrom, RECIPIENT_LIST, msgRoot.as_string())
    smtpObj.quit()
    logger.info('Notification email sent!')

def connect_to_drive():
    credentials_path = get_credentials_path()
    gauth = GoogleAuth()
    # gauth.LoadCredentialsFile("kevin_credentials_storage.txt") # generate token first 
    gauth.LoadCredentialsFile(credentials_path)

    if gauth.credentials is None:
        gauth.LocalWebserverAuth()
    elif gauth.access_token_expired:
        gauth.Refresh()
    else:
        gauth.Authorize()
    # gauth.SaveCredentialsFile("client_secrets.txt") # Must have client_secrets.txt in the same folder
    # gauth.SaveCredentialsFile("kevin_credentials_storage.txt")
    gauth.SaveCredentialsFile(credentials_path)
    
    drive = GoogleDrive(gauth)
    
    logger.info('Connect to Google Drive successful')
    
    return drive
    
def get_sub_folder_id():
    for file_folder in list_parent_folder:
        if file_folder['title'] == yesterday_date:
            sub_folder_id = file_folder['id']
            break     
    logger.info('Get sub-folder id successful')
    
    return sub_folder_id
    
def create_report_directory():
    if not os.path.exists(new_report_path):
        os.makedirs(new_report_path)
    else:
        delete_temp_folder(new_report_path)
        os.makedirs(new_report_path)
        
# Create new report folder in drive
def create_drive_folder():    
    drive = connect_to_drive()
    list_parent_folder = drive.ListFile({'q': "'{}' in parents and trashed=false".format(folder_id)}).GetList()

    # If exist, delete it first
    try:
        for file_folder in list_parent_folder:
            if file_folder['title'] == today_date:
                file_folder.Delete()
                print('Folder ' + today_date + ' deleted')                
    except:
        pass
    
    child_folder = drive.CreateFile({'title': today_date, 'mimeType':'application/vnd.google-apps.folder', 'parents':[{'id':folder_id}]})
    child_folder.Upload()
    print('New report path created')
    
# Delete old drive folder (15 days ago)
def delete_old_drive_folder():
    drive = connect_to_drive()
    del_date = (datetime.now()-timedelta(15)).strftime('%Y%m%d')    
    list_parent_folder = drive.ListFile({'q': "'{}' in parents and trashed=false".format(folder_id)}).GetList()
        
    for file_folder in list_parent_folder:
        if del_date == file_folder['title']:
            file_folder.Delete()
            print('Delete folder: ' + str(file_folder['title'])) 

# Download PN input sheet    
def download_sheet(gsheet_file_id):
    # Will repace if there's an existing folder
    try: 
        # create_report_directory()
        create_drive_folder()
        delete_old_drive_folder()

        logger.info('----------------------------------')
        logger.info('Download sheet..')
        logger.info('Change directory to script path')
        os.chdir(script_path)
        print('Current path: ' + script_path)

        # Download address book
        credentials = get_credentials()
        http = credentials.authorize(httplib2.Http())
        service = build('drive', 'v2', http=http, cache_discovery=False)
        request = service.files().export_media(fileId = gsheet_file_id,
                                            mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # with io.FileIO('PN - User List\PN Input Sheet.xlsx', 'wb') as fh:
        # with io.FileIO(new_report_path + '\PN Input.xlsx', 'wb') as fh:
        with io.FileIO(script_path + '\\' + pn_input_filename + '.xlsx', 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
        
        logger.info('Spreadsheet downloaded')
        
        drive = connect_to_drive()
        sub_folder_id = get_sub_folder_id()
        file_metadata = {'name' : pn_input_filename, 'parents': [ sub_folder_id ]}
        gfile = drive.CreateFile({'title': pn_input_filename, 'mimeType':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', "parents": [{"kind": "drive#fileLink","id": sub_folder_id}]})
        # gfile.SetContentFile(new_report_path)
        gfile.Upload()
        logger.info('Spreadsheet uploaded')

    except Exception as err:
        logger.info('Download gsheet failed')
        logger.info('Error in downloading sheet -> ' + str(err))
        send_message_skype()

#################################################################

def run_custom_sql():
    # DEFINE VARIABLES
    gsheet_url = 'https://docs.google.com/spreadsheets/d/10CHHKhAQWoWiao2t6BEN138h9tWwzWnC__p2Bsi5KOI/edit#gid=1144340668'
    # gsheet_url = 'https://docs.google.com/spreadsheets/d/1kFBUvMlpfTJzOzCp0juCcUey-_SeS1Ln02IC6R2QTCI/edit#gid=170979662'
    gsheet_file_id = (gsheet_url.split('/d/')[1]).split('/')[0]
    excel_main_sheet = None
    gsheet_credential = os.getcwd() + '\PN - User List\KEVIN_CLIENT_SERVICE.json'
    gc = pygsheets.authorize(service_file=gsheet_credential)
    
    # Download sheet to excel file
    download_sheet(gsheet_file_id)

    # Read PN Input excel file
    logger.info('Opening ' + pn_input_filename + '.xlsx..')
    wb = load_workbook(script_path + '\\' + pn_input_filename + '.xlsx')
    # wb = load_workbook(new_report_path + '\PN Input.xlsx')
    sheet_names = wb.sheetnames

    num_of_case = 0
    for a in range(len(sheet_names)):
        if 'Case' in sheet_names[a]:
            num_of_case = num_of_case + 1
    print('Num of cases: ' + str(num_of_case))

    for x in range(num_of_case):
        try:
            case_sheet = 'Case ' + str(int(x+1))
            excel_case_sheet = wb[case_sheet]

            # Check for null value                  
            if not excel_case_sheet['A2'].value == None:
                excel_main_sheet = case_sheet
                logger.info('Value detected - use sheet: ' + excel_main_sheet)
            else:
                logger.info('No value detected on sheet: ' + case_sheet)
                continue

            logger.info('Reading Data From Excel...')
            # sheet = wb.get_sheet_by_name(excel_main_sheet)
            sheet = wb[excel_main_sheet]
            
            ####################################################

            # Initialize start row
            current_row = 2
            # str_category = None
            # num_category = 0

            logger.info('===================================')
            logger.info('Get custom SQL..')

            try:
                # Case 1
                if excel_main_sheet == 'Case 1':
                    logger.info('USING CASE 1')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        category = '\'' + sheet['B' + str(current_row)].value + '\''
                        last_login = int(sheet['C' + str(current_row)].value)

                        try:
                            str_category = category.replace(', ', '\',\'')
                            str_category = category.replace(',', '\',\'')
                        except:
                            pass

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Category: ' + str_category)
                        logger.info('Last Login: ' + str(last_login))

                        custom_sql = '''
                                select distinct
                                    concat('user_id=',cast(u.userid as varchar)) list_userid
                                    from user_profile as u
                                    join order_mart__order_item_profile as p
                                    on u.userid = p.userid
                                    where
                                    p.main_category in (%s)
                                    and
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and
                                    u.is_buyer = 1
                                ''' % (str_category, last_login)
                        
                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 2
                elif excel_main_sheet == 'Case 2':
                    logger.info('USING CASE 2')
                    for x in range(300):        
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        last_login = int(sheet['B' + str(current_row)].value)
                        date_1 = sheet['C' + str(current_row)].value
                        date_1 = datetime.strftime(date_1, '%Y-%m-%d')
                        date_2 = sheet['D' + str(current_row)].value
                        date_2 = datetime.strftime(date_2, '%Y-%m-%d')

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Last Login: ' + str(last_login))
                        logger.info('Date 1: ' + str(date_1))
                        logger.info('Date 2: ' + str(date_2))

                        custom_sql = ''' 
                                with a as(
                                select distinct u.userid
                                    from user_profile as u
                                    where
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                ),

                                b as
                                (
                                select distinct
                                userid
                                from order_mart__order_item_profile
                                where grass_date between date'%s' and date'%s'
                                )

                                select distinct
                                concat('user_id=',cast(a.userid as varchar)) list_userid
                                from a      
                                left join b
                                on a.userid = b.userid
                                where b.userid is null
                                ''' % (last_login, date_1, date_2)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 3
                elif excel_main_sheet == 'Case 3':
                    logger.info('USING CASE 3')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        category = '\'' + sheet['B' + str(current_row)].value + '\''
                        last_login = int(sheet['C' + str(current_row)].value)
                        date_1 = sheet['D' + str(current_row)].value
                        date_1 = datetime.strftime(date_1, '%Y-%m-%d')
                        date_2 = sheet['E' + str(current_row)].value
                        date_2 = datetime.strftime(date_2, '%Y-%m-%d')

                        try:
                            str_category = category.replace(', ', '\',\'')
                            str_category = category.replace(',', '\',\'')
                        except:
                            pass

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Category: ' + str_category)
                        logger.info('Last Login: ' + str(last_login))
                        logger.info('Date 1: ' + str(date_1))
                        logger.info('Date 2: ' + str(date_2))

                        custom_sql = '''
                                with a as(
                                select distinct
                                    u.userid
                                    from user_profile as u
                                    join order_mart__order_item_profile as p
                                    on u.userid = p.userid
                                    where
                                    p.main_category in (%s)
                                    and
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and
                                    is_buyer = 1
                                ),

                                --shopped for the last 30 days
                                b as
                                (
                                select distinct
                                userid
                                from order_mart__order_item_profile
                                where grass_date between date'%s' and date'%s'
                                )

                                --final
                                select distinct
                                concat('user_id=',cast(a.userid as varchar)) list_userid
                                from a
                                left join b
                                on a.userid = b.userid
                                where b.userid is null
                                ''' % (str_category, last_login, date_1, date_2)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 4 - Certain category buyers who haven't shopped other main category
                elif excel_main_sheet == 'Case 4':
                    logger.info('USING CASE 4')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        category = '\'' + sheet['B' + str(current_row)].value + '\''
                        last_login = int(sheet['C' + str(current_row)].value)
                        category_not_shopped = '\'' + sheet['D' + str(current_row)].value + '\''

                        try:
                            str_category = category.replace(', ', '\',\'')
                            str_category = category.replace(',', '\',\'')
                        except:
                            pass

                        try:
                            str_category_not_shopped = category_not_shopped.replace(', ', '\',\'')
                            str_category_not_shopped = category_not_shopped.replace(',', '\',\'')
                        except:
                            pass

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Category: ' + str_category)
                        logger.info('Last Login: ' + str(last_login))
                        logger.info('Category(Haven\'t shopped): ' + str_category_not_shopped)

                        custom_sql = '''
                                with a as (
                                select distinct
                                    u.userid
                                    from user_profile as u
                                    join order_mart__order_item_profile as p
                                    on u.userid = p.userid
                                    where
                                    p.main_category in (%s)
                                    and
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and
                                    is_buyer = 1
                                ),

                                --shopped certain category
                                b as
                                (
                                select distinct
                                userid
                                from order_mart__order_item_profile
                                where 
                                main_category in (%s)
                                )

                                --final
                                select distinct
                                concat('user_id=',cast(a.userid as varchar)) list_userid
                                from a
                                left join b
                                on a.userid = b.userid
                                where b.userid is null
                                ''' % (str_category, last_login, str_category_not_shopped)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 5 - Certain category buyers who haven't shopped other sub-category
                elif excel_main_sheet == 'Case 5':
                    logger.info('USING CASE 5')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        category = '\'' + sheet['B' + str(current_row)].value + '\''
                        last_login = int(sheet['C' + str(current_row)].value)
                        subcategory_not_shopped = '\'' + sheet['D' + str(current_row)].value + '\''

                        try:
                            str_category = category.replace(', ', '\',\'')
                            str_category = category.replace(',', '\',\'')
                        except:
                            pass

                        try:
                            str_subcategory_not_shopped = subcategory_not_shopped.replace(', ', '\',\'')
                            str_subcategory_not_shopped = subcategory_not_shopped.replace(',', '\',\'')
                        except:
                            pass

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Category: ' + str_category)
                        logger.info('Last Login: ' + str(last_login))
                        logger.info('Sub-category(Haven\'t shopped): ' + str_subcategory_not_shopped)
                        
                        custom_sql = '''
                                with a as (
                                select distinct
                                    u.userid
                                    from user_profile as u
                                    join order_mart__order_item_profile as p
                                    on u.userid = p.userid
                                    where
                                    p.main_category in (%s)
                                    and
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and
                                    is_buyer = 1
                                ),

                                --shopped certain category
                                b as
                                (
                                select distinct
                                userid
                                from order_mart__order_item_profile
                                where 
                                sub_category in (%s)
                                )

                                --final
                                select distinct
                                concat('user_id=',cast(a.userid as varchar)) list_userid
                                from a
                                left join b
                                on a.userid = b.userid
                                where b.userid is null
                                ''' % (str_category, last_login, str_subcategory_not_shopped)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 6
                elif excel_main_sheet == 'Case 6':
                    logger.info('USING CASE 6')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        last_login = int(sheet['B' + str(current_row)].value)
                        gender = sheet['C' + str(current_row)].value

                        str_gender = None
                        if str(gender.lower()) == 'male':
                            str_gender = '1,3'
                        elif str(gender.lower()) == 'female':
                            str_gender = '2,4'
                
                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Last Login: ' + str(last_login))
                        logger.info('Gender: ' + str_gender)

                        custom_sql = '''
                                select distinct
                                    concat('user_id=',cast(u.userid as varchar)) list_userid
                                    from user_profile as u
                                    join order_mart__order_item_profile as p
                                    on u.userid = p.userid 
                                    where
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and
                                    u.gender in (%s)
                                ''' % (last_login, str_gender)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 7
                elif excel_main_sheet == 'Case 7':
                    logger.info('USING CASE 7')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        category = '\'' + sheet['B' + str(current_row)].value + '\''
                        last_login = int(sheet['C' + str(current_row)].value)
                        gender = sheet['D' + str(current_row)].value

                        try:
                            str_category = category.replace(', ', '\',\'')
                            str_category = category.replace(',', '\',\'')
                        except:
                            pass

                        str_gender = None
                        if str(gender.lower()) == 'male':
                            str_gender = '1,3'
                        elif str(gender.lower()) == 'female':
                            str_gender = '2,4'
                
                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Category: ' + str_category)
                        logger.info('Last Login: ' + str(last_login))
                        logger.info('Gender: ' + str_gender)
                                            
                        custom_sql = '''
                                select distinct
                                    concat('user_id=',cast(u.userid as varchar)) list_userid
                                    from user_profile as u
                                    join order_mart__order_item_profile as p
                                    on u.userid = p.userid
                                    where
                                    p.main_category in (%s)
                                    and
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and
                                    u.gender in (%s)
                                ''' % (str_category, last_login, str_gender)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break
                
                # Case 8
                elif excel_main_sheet == 'Case 8':
                    logger.info('USING CASE 8')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        sub_category = '\'' + sheet['B' + str(current_row)].value + '\''
                        last_login = int(sheet['C' + str(current_row)].value)

                        try:
                            str_sub_category = sub_category.replace(', ', '\',\'')
                            str_sub_category = sub_category.replace(',', '\',\'')
                        except:
                            pass

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Sub-category: ' + str_sub_category)
                        logger.info('Last Login: ' + str(last_login))

                        custom_sql = '''
                                select distinct
                                    concat('user_id=',cast(u.userid as varchar)) list_userid
                                    from user_profile as u
                                    join order_mart__order_item_profile as p
                                    on u.userid = p.userid
                                    where
                                    p.sub_category in (%s)
                                    and
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and
                                    u.is_buyer = 1
                                ''' % (str_sub_category, last_login)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break 

                # Case 9
                elif excel_main_sheet == 'Case 9':
                    logger.info('USING CASE 9')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        sub_category = '\'' + sheet['B' + str(current_row)].value + '\''
                        last_login = int(sheet['C' + str(current_row)].value) 
                        date_1 = sheet['D' + str(current_row)].value
                        date_1 = datetime.strftime(date_1, '%Y-%m-%d')
                        date_2 = sheet['E' + str(current_row)].value
                        date_2 = datetime.strftime(date_2, '%Y-%m-%d')

                        try:
                            str_sub_category = sub_category.replace(', ', '\',\'')
                            str_sub_category = sub_category.replace(',', '\',\'')
                        except:
                            pass

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Sub-category: ' + str_sub_category)
                        logger.info('Last Login: ' + str(last_login))
                        logger.info('Date 1: ' + str(date_1))
                        logger.info('Date 2: ' + str(date_2))

                        custom_sql = '''
                                with a as(
                                select distinct
                                    u.userid
                                    from user_profile as u
                                    join order_mart__order_item_profile as p
                                    on u.userid = p.userid
                                    where
                                    p.sub_category in (%s)
                                    and
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and
                                    u.is_buyer = 1
                                ),

                                --shopped for the last 30 days
                                b as
                                (
                                select distinct
                                userid
                                from order_mart__order_item_profile
                                where grass_date between date'%s' and date'%s'
                                )

                                --final
                                select distinct
                                concat('user_id=',cast(a.userid as varchar)) list_userid
                                from a
                                left join b
                                on a.userid = b.userid
                                where b.userid is null
                                ''' % (str_sub_category, last_login, date_1, date_2)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break 
                
                # Case 10
                elif excel_main_sheet == 'Case 10':
                    logger.info('USING CASE 10')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        shop_id = int(sheet['B' + str(current_row)].value)

                        try:
                            str_category = category.replace(', ', '\',\'')
                            str_category = category.replace(',', '\',\'')
                        except:
                            pass

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Shop ID: ' + str(shop_id))

                        custom_sql = '''
                                select
                                'user_id='||cast(userid as varchar) as list_userid
                                from shopee_follower_id_db__shop_follow_tab
                                where shopid in (%s) 
                                and status = 1
                                ''' % (shop_id)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break 
                            
                ################################## DIGITAL PRODUCTS

                # Case 11
                elif excel_main_sheet == 'Case 11':
                    logger.info('USING CASE 11')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        last_login = int(sheet['B' + str(current_row)].value)

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Last Login: ' + str(last_login))

                        custom_sql = '''    
                                with a as(
                                select distinct
                                    u.userid
                                    from user_profile as u
                                    where
                                    last_login >= current_date - interval '%s' day
                                    and
                                    u.status = 1
                                    and 
                                    u.is_buyer=1
                                ),

                                --DP Buyers
                                b as(
                                select distinct o.user_id as userid
                                from shopee_digital_product_order_id_db__order_tab AS o
                                )

                                select distinct 
                                concat('user_id=',cast(a.userid as varchar)) as list_userid
                                from a
                                join b 
                                on a.userid = b.userid
                                ''' % (last_login)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 12 - Users that login within the last certain days who haven't shopped digital products	
                elif excel_main_sheet == 'Case 12':
                    logger.info('USING CASE 12')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        last_login = int(sheet['B' + str(current_row)].value)

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Last Login: ' + str(last_login))

                        custom_sql = '''
                                with b as(
                                    select distinct o.user_id as userid
                                    from shopee_digital_product_order_id_db__order_tab AS o
                                )

                                select distinct 
                                    concat('user_id=',cast(u.userid as varchar)) list_userid
                                    from user_profile as u 
                                    where
                                    last_login >= current_date - interval '%s' day
                                    and 
                                    u.userid not in (select distinct userid from b)
                                    and 
                                    u.status = 1
                                ''' % (last_login)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 13 - Certain category buyers that login within the last certain days who haven't shopped digital products	
                elif excel_main_sheet == 'Case 13':
                    logger.info('USING CASE 13')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        category = '\'' + sheet['B' + str(current_row)].value + '\''
                        last_login = int(sheet['C' + str(current_row)].value)

                        try:
                            str_category = category.replace(', ', '\',\'')
                            str_category = category.replace(',', '\',\'')
                        except:
                            pass

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Main category: ' + str(str_category))
                        logger.info('Last Login: ' + str(last_login))

                        custom_sql = '''
                                with b as(
                                    select distinct o.user_id as userid
                                    from shopee_digital_product_order_id_db__order_tab AS o
                                )

                                select distinct 
                                    concat('user_id=',cast(u.userid as varchar)) list_userid
                                    from user_profile as u 
                                    join order_mart__order_item_profile as p 
                                    on u.userid = p.userid
                                    where
                                    p.main_category in (%s) 
                                    and
                                    last_login >= current_date - interval '%s' day
                                    and 
                                    u.userid not in (select distinct userid from b)
                                    and 
                                    u.status = 1
                                ''' % (str_category, last_login)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 14 - USER FILTERED BY ITEMID
                elif excel_main_sheet == 'Case 14':
                    logger.info('USING CASE 14')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        item_id = int(sheet['B' + str(current_row)].value)
                        date_1 = sheet['C' + str(current_row)].value
                        date_1 = datetime.strftime(date_1, '%Y-%m-%d')
                        date_2 = sheet['D' + str(current_row)].value
                        date_2 = datetime.strftime(date_2, '%Y-%m-%d')

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Item ID: ' + str(item_id))
                        logger.info('Date 1: ' + str(date_1))
                        logger.info('Date 2: ' + str(date_2))

                        custom_sql = '''
                                select distinct
                                concat('user_id=',cast(p.userid as varchar)) as list_userid
                                from order_mart__order_item_profile as p
                                where itemid in (%s)
                                and grass_date between date'%s' and date'%s'
                                ''' % (item_id, date_1, date_2)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break
                
                # Case 15 - GOYANG SHOPEE USER
                elif excel_main_sheet == 'Case 15':
                    logger.info('USING CASE 15')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        date_1 = sheet['B' + str(current_row)].value
                        date_1 = datetime.strftime(date_1, '%Y-%m-%d')
                        date_2 = sheet['C' + str(current_row)].value
                        date_2 = datetime.strftime(date_2, '%Y-%m-%d')

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Date 1: ' + str(date_1))
                        logger.info('Date 2: ' + str(date_2))

                        custom_sql = '''
                                select distinct
                                concat('user_id=',cast(a1.userid as varchar)) as list_userid
                                from user_login_record_id_db__user_login_record_tab as a1
                                where a1.grass_date between date'{}' and date'{}'
                                and a1.userid in (select distinct userid from shopee_coins_id_db__coin_transaction_tab where lower(from_utf8(from_base64(info))) like '%goyang shopee%')
                                '''.format(date_1, date_2)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 16 - QUIZ SHOPEE USER
                elif excel_main_sheet == 'Case 16':
                    logger.info('USING CASE 16')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        date_1 = sheet['B' + str(current_row)].value
                        date_1 = datetime.strftime(date_1, '%Y-%m-%d')
                        date_2 = sheet['C' + str(current_row)].value
                        date_2 = datetime.strftime(date_2, '%Y-%m-%d')

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Date 1: ' + str(date_1))
                        logger.info('Date 2: ' + str(date_2))
                    
                        custom_sql = '''
                                select distinct
                                concat('user_id=',cast(a.userid as varchar)) as list_userid
                                from shopee_id_bi_team__kuis_players as a
                                join shopee_gamehq_id_db__hq_sessions_tab as b 
                                on a.session_id = cast(b.id as varchar)
                                join shopee_gamehq_id_db__hq_events_tab as c 
                                on trim(a.event_id) = cast(c.id as varchar)
                                where date(from_unixtime(b.start,'Asia/Jakarta')) >= date('%s')
                                and date(from_unixtime(b.start,'Asia/Jakarta')) <= date('%s')
                                ''' % (date_1, date_2)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 17 - ONE DOLLAR GAME USER
                elif excel_main_sheet == 'Case 17':
                    logger.info('USING CASE 17')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        date_1 = sheet['B' + str(current_row)].value
                        date_1 = datetime.strftime(date_1, '%Y-%m-%d')
                        date_2 = sheet['C' + str(current_row)].value
                        date_2 = datetime.strftime(date_2, '%Y-%m-%d')

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Date 1: ' + str(date_1))
                        logger.info('Date 2: ' + str(date_2))

                        custom_sql = '''
                                select distinct
                                concat('user_id=',cast(p.userid as varchar)) as list_userid, grass_date
                                from order_mart__order_item_profile as p
                                where bi_excluded = 'OneDollarGame'
                                and grass_date between date'%s' and date'%s'
                                ''' % (date_1, date_2)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break

                # Case 18 - DAILY PRIZE USER
                elif excel_main_sheet == 'Case 18':
                    logger.info('USING CASE 18')
                    for x in range(300):
                        sql_file_name = sheet['A' + str(current_row)].value # Max 25 characters
                        date_1 = sheet['B' + str(current_row)].value
                        date_1 = datetime.strftime(date_1, '%Y-%m-%d')
                        date_2 = sheet['C' + str(current_row)].value
                        date_2 = datetime.strftime(date_2, '%Y-%m-%d')

                        logger.info('----------------------------------')
                        logger.info('Row ' + str(int(current_row-1)))
                        logger.info('Filename: ' + sql_file_name)
                        logger.info('Date 1: ' + str(date_1))
                        logger.info('Date 2: ' + str(date_2))

                        custom_sql = '''
                                select distinct
                                concat('user_id=',cast(dp.user_id as varchar)) as list_userid
                                from shopee_id_mk_team__daily_prize_vfive as dp
                                left join user_profile as u 
                                on u.userid = cast(dp.user_id as int)
                                where date(date_parse(cast(dp.datte as varchar), '%Y-%m-%d %H:%i:%s')) >= date('{}')
                                and date(date_parse(cast(dp.datte as varchar), '%Y-%m-%d %H:%i:%s')) <= date('{}')                                
                                '''.format(date_1, date_2)

                        logger.info('Query:' + custom_sql)
                        # Call run_csv_sql function
                        run_csv_sql(custom_sql, sql_file_name)

                        current_row = current_row + 1
                        # If the next row has empty value, break the loop
                        if sheet['A' + str(current_row)].value == None:
                            break
                else:
                    logger.info('Case sheet incorrect')

            except Exception as err:
                logger.info('Error in run_custom_sql() -> ' + str(err))
                send_message_skype()
                pass         

            # Open Google Sheet
            gsheet_main_sheet = excel_main_sheet
            wb_gsheet = gc.open_by_url(gsheet_url)
            sheet_gsheet = wb_gsheet.worksheet_by_title(gsheet_main_sheet)

            # Deleting data in Google Sheet
            logger.info('----------------------------------')
            logger.info('Delete all data in sheet: ' + str(gsheet_main_sheet))
            sheet_gsheet.clear('A2', end=None)       

        except Exception as err: 
            logger.info(str(err))

# Run custom sql and generate it to CSV
def run_csv_sql(custom_sql, sql_file_name):
    try:
        logger.info('----------------------------------')
        logger.info('Start Running query ' + sql_file_name)
        [col_name, rows] = run_presto(custom_sql) # Run custom sql depends what case the user wants
        logger.info('Running '+ sql_file_name + ' Done')
        csv_file_name = sql_file_name.replace('.sql', '_')

        logger.info("CSV file path: " + csv_file_name)

        df = pd.DataFrame(rows, columns=col_name) # convert the query results into a dataframe
        
        logger.info('Starting save the query results')
        logger.info('Total Rows: ' + str(len(df.index)))
        NUMBER_OF_SPLITS = math.ceil(len(df.index)/1000000.00)
        logger.info('Will split into : ' + str(NUMBER_OF_SPLITS) + ' files')
        
        # Creating txt file
        if len(df.index) == 0:
            txt_file = open(new_report_path + '\Troubleshooting Mismatched Input.txt', 'a') 
            txt_file.write('Wrong classification for {}, please check your input again'.format(sql_file_name)) 
            txt_file.write('\n')
            txt_file.close() 
            logger.info('Mismatched Input! - Write to txt file successful')

        sub_folder_id = get_sub_folder_id() # Get sub folder ID
        
        # Upload CSV to Google Drive
        for i, new_df in enumerate(np.array_split(df, NUMBER_OF_SPLITS)):
            # fill_email_excel(NUMBER_OF_SPLITS, i)
            file_metadata = {'name' : csv_file_name + '_' + str(i+1), 'parents': [ sub_folder_id ]}
            gfile = drive.CreateFile({'title': csv_file_name + '_' + str(i+1), 'mimeType':'text/csv', "parents": [{"kind": "drive#fileLink","id": sub_folder_id}]})
            # gfile.SetContentFile(new_report_path)
            gfile.Upload()
            logger.info('Save query results as ' + csv_file_name + '_' + str(i+1) +'.csv')
            logger.info(', with total : ' + str(len(new_df.index)) + ' rows')
            
        logger.info('CSV uploaded')
    except Exception as err:
        logger.info('Error in run_csv_sql -> ' + str(err))
        send_message_skype()

if __name__ == '__main__':

    # Download PN Input sheet
    # download_sheet()
    connect_to_drive()
    # Get custom sql and run sql to csv
    run_custom_sql()

    # Run sql to csv
    # run_csv_sql(custom_sql, sql_file_name)

    # Zip the CSV file
    # zip_file()

