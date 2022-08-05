# -*- coding: utf-8 -*-
import smtplib
from email.MIMEMultipart import MIMEMultipart
from email.MIMEText import MIMEText
from email.MIMEImage import MIMEImage
from email.MIMEBase import MIMEBase
from email import encoders
import shutil
import httplib2
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.discovery import build
import oauth2client
from oauth2client import file
from oauth2client import client
from oauth2client import tools
import io
import os, os.path
import datetime
from datetime import date, timedelta
from openpyxl import load_workbook
import excel2img
from shutil import copyfile
import xlrd
import logging
import sys
import zipfile
from zipfile import ZipFile
########################################## CONFIGURATION REQUIRED ##########################################

### AUTO EMAIL CONFIGURATION TEMPLATE
AUTO_EMAIL_TEMPLATE_PATH = 'G:\My Drive\AUTOMATED_REPORTS\PN - Automation\PN - User List\\'
AUTO_EMAIL_TEMPLATE_NAME = 'PN - Automation Email' # xlsx email template file
AUTO_EMAIL_MAIN_SHEET = "Config"

### REPORT RECIPIENT LIST (GOOGLE SHEET ADDRESS BOOK)
ADDRESS_BK_URL = 'https://docs.google.com/spreadsheets/d/1D_eGQ4CpRx4tHp6BmM0Z29HRUCLXuwfGUJvAHoutwkU/edit#gid=1915221268'
ADDRESS_BK_URL_FILE_ID = (ADDRESS_BK_URL.split('/d/')[1]).split('/')[0]
ADDRESS_BK_SHEETNAME = 'PN - Automation'
CLIENT_SECRET_FILE = '' # need to put under same directory as this .py file  ## erased for confidentiality

### TESTING EMAIL RECIPIENT LIST (LEAVE IT BLANK IF YOU WANT TO SEND ACTUAL EMAIL)
# TEST_RECIPIENT_LIST = ['chianyi.goh@shopee.com']
TEST_RECIPIENT_LIST = []

### EMAIL CREDENTIALS
SENDER = 'Kevin Kurnia'
LOGIN_EMAIL = '' ## erased for confidentiality
LOGIN_PASSWORD = '' ## erased for confidentiality
PERSONAL_EMAIL = 'kevin.kurnia@shopee.com'

# EMAIL_SUBJECT = 'Seller Management - Weekly RM Report [%s] - %s' # can include variable(s) %s if different title for different emails
# EMAIL_SUBJECT = 'Seller Management - Weekly RM Report' + ' ' + REPORT_DATE # or fix the title for all emails
# EMAIL_MESSAGE: can include variable(s) %s if different message for different emails

#############################################################################################################

### FOR WRITING LOG FILE
log_formatter = '%(asctime)s - %(levelname)s - %(message)s'
logging.basicConfig(level=logging.INFO, format=log_formatter)
logger = logging.getLogger(__name__)
handler = logging.FileHandler(AUTO_EMAIL_TEMPLATE_PATH + AUTO_EMAIL_TEMPLATE_NAME + '.log')
handler.setLevel(logging.INFO)
handler.setFormatter(logging.Formatter(log_formatter))
logger.addHandler(handler)

# download email address from googlesheet
SCOPES = 'https://www.googleapis.com/auth/drive'
APPLICATION_NAME = 'Google drive API Python Quickstart'

# for download address book
def get_credentials():
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir, 'auto_download.json')
    store = oauth2client.file.Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        credentials = tools.run_flow(flow, store)
        logger.info('Storing credentials to ' + credential_path)
    return credentials

def delete_address_book():
    # delete address book
    dir_path = os.path.dirname(os.path.realpath(__file__))
    if empty_list(TEST_RECIPIENT_LIST) == True:
        add_bk = dir_path + '\Email Address.xlsx'
        if os.path.isfile(add_bk):
            os.remove(add_bk)
        else:  ## Show an error ##
            logger.info("Error: %s file not found" % add_bk)

def delete_temp_folder(FolderPath):
    try:
        shutil.rmtree(FolderPath)
    except OSError, e:  ## if failed, report it back to the user ##
        logger.info ("Error: %s - %s." % (e.filename, e.strerror))

def prepare_email_folder(path,filename,main_sheet):

    logger.info('Preparing Email Sending Folders....')
    rpt_sht_img_count = {}

    email_template_full_path = path + filename + '.xlsx' # Email Template
    wb = load_workbook(email_template_full_path, data_only=True)
    esheet = wb[main_sheet] # Configuration Page
    enum = esheet.max_row
    # print(enum)
    logger.info("Banyak enum: " + str(enum))

    # report_date = esheet['B3'].value # after formatted
    report_date = (date.today()).strftime('%Y%m%d')
    logger.info('Report Date' + str(report_date))

    report_dir = esheet['B4'].value # with '%d'
    report_dir = report_dir.replace('%d',report_date)
    logger.info("Opening " + report_dir)
    new_report_folder = report_dir + "Email\\" # Open a new folder /Email/

    if not os.path.exists(new_report_folder):
        os.makedirs(new_report_folder)
    else:
        delete_temp_folder(new_report_folder)
        os.makedirs(new_report_folder)

    for i in range(7,enum+1):
        email_group = str(esheet['A' + str(i)].value) # 1 email_group = 1 folder
        print(email_group)
        report_file = str(esheet['B' + str(i)].value)
        
        report_file = report_file.replace('%d', report_date)

        attach_screenshot = esheet['C' + str(i)].value # A or S or A+S

        report_file_full_path = report_dir + report_file
        output_folder = new_report_folder + email_group + "\\" #jovita

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        if attach_screenshot != 'S': # A or A+S
            copyfile(report_file_full_path, output_folder + report_file) #copy the report to the new folder

        if attach_screenshot != 'A': # S or A+S

            sheet_ranges = esheet['D' + str(i)].value
            sheet_range_list = sheet_ranges.split(",")

            for (j,sheet_range) in enumerate(sheet_range_list):

                report_sheet = sheet_range.split("[")[0] # sheet_name
                report_range_list = ((sheet_range.split("[")[1]).split("]")[0]).split("/") # range list to screenshot for this sheet
                rpt_sht_img_count['Rpt_' + str(i-7+1) + '_Sht' + str(j+1)] = len(report_range_list)

                for (k,image_range) in enumerate(report_range_list):
                    image_file = output_folder + "Rpt_" + str(i-7+1) + "_Sht" + str(j+1) + "_Img" + str(k+1) + ".png"

                    if len(image_range) == 0:
                        excel2img.export_img(report_file_full_path, image_file, report_sheet)
                    else:
                        excel2img.export_img(report_file_full_path, image_file, report_sheet, image_range)

    return [new_report_folder,report_date]

def attach_image(emailImage,dir,ImagePng):

    fp = open(dir + "\\" + ImagePng, 'rb')
    msgImage = MIMEImage(fp.read())
    msgImage.add_header('Content-ID','<' + ImagePng.replace(".png", "") + '>')
    msgRoot.attach(msgImage)
    fp.close()

    emailImage = emailImage + "<img src='cid:" + ImagePng.replace(".png", "") + "'" + " style='margin:0; padding:0'><br><br>"

    return emailImage

# PREPARE EXCEL ATTACHMENT
def attach_report_files(dir,report_attach):

    fp = open(os.path.join(dir,report_attach), 'rb')
    excel_file = MIMEBase('application', 'vnd.ms-excel')
    excel_file.set_payload(fp.read())
    fp.close()
    encoders.encode_base64(excel_file)
    excel_file.add_header('Content-Disposition', 'attachment', filename=report_attach)
    msgRoot.attach(excel_file)

def empty_list(input_list):
    sum = 0
    for list in input_list:
        sum += len(list)
    if sum == 0:
        return True
    else:
        return False

def download_address_book():
    ## downloading address book
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = build('drive', 'v2', http=http, cache_discovery=False)

    request = service.files().export_media(fileId = ADDRESS_BK_URL_FILE_ID,
                                           mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    with io.FileIO('PN - User List\Email Address.xlsx', 'wb') as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            # logger.info("Email List Download %d%%." % int(status.progress() * 100))

def get_recipient_list(test_recipient, address_header):
    # copy email address in googlesheet into python list recipients
    RECIPIENT_LIST = []
    CC_RECIPIENT_LIST = []
    BCC_RECIPIENT_LIST = []
    if empty_list(test_recipient) == True:
        logger.info('Going to send email to ALL recipients!!')
        matched_version = 0

        # Download the Gdoc Address Book
        download_address_book()

        ## load EMAIL ADDRESS workbook
        # logger.info(os.getcwd())
        with xlrd.open_workbook('PN - User List\Email Address.xlsx') as workbook:
            try:
                worksheet = workbook.sheet_by_name(ADDRESS_BK_SHEETNAME)
            except xlrd.XLRDError:
                logger.error('No sheet named "<' + ADDRESS_BK_SHEETNAME +'>". Please double check ADDRESS_BK_SHEETNAME or Sheet Name in Gdoc.')
                sys.exit(1)
            except Exception as e:
                logger.info(str(e))

            ncol = worksheet.ncols
            if ncol==1: #only one version to send out
                nrow = worksheet.nrows
                for j in range(1, nrow):
                    address = worksheet.cell_value(j, 0).encode("ascii", "ignore")
                    address = address.replace(" ", "")
                    if len(address) == 0:
                        continue
                    else:
                        if str(address).startswith('[CC]'):
                            CC_RECIPIENT_LIST.append(str(address).split('[CC]')[1])
                        elif str(address).startswith('[BCC]'):
                            BCC_RECIPIENT_LIST.append(str(address).split('[BCC]')[1])
                        else:
                            RECIPIENT_LIST.append(str(address))

            elif ncol>1:
                for k in range(0, ncol):
                    title = worksheet.cell_value(0, k).encode("ascii", "ignore")

                    # remove unnecessary spaces in the end to prevent unmatched
                    while title[-1] == " ":
                        title = title[:-1]
                    while address_header[-1] == " ":
                        address_header = address_header[:-1]

                    if title == address_header:
                        matched_version = 1
                        curr_col = k
                        break

                if matched_version == 0:
                    logger.error('No Email Version named "<' + address_header + '>" in Gdoc Address Book". Please double check your Email Version Title.')
                    sys.exit(1)

                nrow = worksheet.nrows
                for j in range(1, nrow):
                    address = worksheet.cell_value(j, curr_col).encode("ascii", "ignore")
                    address = address.replace(" ", "")
                    if len(address) == 0:
                        continue
                    else:
                        if str(address).startswith('[CC]'):
                            CC_RECIPIENT_LIST.append(str(address).split('[CC]')[1])
                        elif str(address).startswith('[BCC]'):
                            BCC_RECIPIENT_LIST.append(str(address).split('[BCC]')[1])
                        else:
                            RECIPIENT_LIST.append(str(address))
    else:
        logger.info('Going to send email to TEST recipient(s)!!')
        RECIPIENT_LIST = test_recipient
    return [RECIPIENT_LIST, CC_RECIPIENT_LIST, BCC_RECIPIENT_LIST]

def send_email(emailImage,recipient_list, cc_recipient_list, bcc_recipient_list, email_subject,email_msg):
    # prepare to send out email!
    strFrom = SENDER + ' ' + PERSONAL_EMAIL
    msgRoot['Subject'] = email_subject
    msgRoot['From'] = strFrom

    msgRoot['To'] = ", ".join(recipient_list)
    if len(cc_recipient_list) > 0:
        msgRoot['CC'] = ", ".join(cc_recipient_list)

    To_Addrs = recipient_list + cc_recipient_list + bcc_recipient_list

    msgText = MIMEText(
        email_msg + emailImage + '</i>', 'html')
    msgRoot.attach(msgText)

    smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
    smtpObj.ehlo()
    smtpObj.starttls()
    smtpObj.login(LOGIN_EMAIL, LOGIN_PASSWORD)
    smtpObj.sendmail(strFrom, To_Addrs, msgRoot.as_string())
    smtpObj.quit()
    logger.info("Email successfully sent")

if __name__ == '__main__':
    # Open Excel Email Template to prepare the attachments and screenshots
    [REPORT_EMAIL_FOLDER, REPORT_DATE] = prepare_email_folder(AUTO_EMAIL_TEMPLATE_PATH, AUTO_EMAIL_TEMPLATE_NAME, AUTO_EMAIL_MAIN_SHEET)
    REPORT_DIR = 'G:/My Drive/AUTOMATED_REPORTS/PN - Monthlong Cashback Weekly Reminder/' + REPORT_DATE
    REPORT_DATE2 = datetime.datetime.today().strftime('%Y%m%d')
    LINK_REPORT = 'https://drive.google.com/open?id=1VHqUSXMeVIp6q17uk8xuZ6mcjpqw2WL1'

    EMAIL_SUBJECT = 'PN Reminder User List ' + REPORT_DATE2
    EMAIL_MESSAGE = '''  
            Hi Comms Team, <br><br>
            Pls find the user list for month-long cashback PN reminder which will be sent every Saturday. <br>
            Link: %s <br><br>
            Thank you.
            ''' % (LINK_REPORT)
            
    # Loop for every folder/file under REPORT_EMAIL_FOLDER to send out email
    for root, dirs, files in os.walk(REPORT_EMAIL_FOLDER): 
        for rpt_folder in dirs: # Email Group/Version # Slain ALL ada apa lg
            logger.info('Sending Email for ' + rpt_folder)
            RECIPIENT_LISTS = get_recipient_list(TEST_RECIPIENT_LIST, rpt_folder)

            emailImage = '<br>'
            msgRoot = MIMEMultipart()
            RECIPIENT_LIST = []
            CURR_REPORT_DIR = os.path.join(root, rpt_folder)

            # os.chdir(CURR_REPORT_DIR)

            # logger.info(CURR_REPORT_DIR)
            # logger.info(os.getcwd())

            # with ZipFile(CURR_REPORT_DIR + '/pn_user_list.zip', 'w', zipfile.ZIP_DEFLATED) as zipObj:
            #     for filename in os.listdir(CURR_REPORT_DIR):
            #         if filename.endswith(".csv"):
            #             logger.info('Zipping: ' + filename)
            #             # zipObj.write(CURR_REPORT_DIR + '/' + filename)
            #             zipObj.write(filename)
            #         else:
            #             continue
                
            # if filename.endswith(".png"): # image to attach
            #     emailImage = attach_image(emailImage, CURR_REPORT_DIR, filename)
            # elif filename.endswith(".zip"):
            #     attach_report_files(CURR_REPORT_DIR,filename)
            # elif (not filename.endswith(".png") and filename.endswith(".xlsx")):
            #     attach_report_files(CURR_REPORT_DIR,filename)
            # elif filename.endswith(".csv"):
            #     attach_report_files(CURR_REPORT_DIR,filename)
            # else:
            #     continue

            # send_email(emailImage,RECIPIENT_LISTS[0],RECIPIENT_LISTS[1],RECIPIENT_LISTS[2], EMAIL_SUBJECT%(rpt_folder,REPORT_DATE), EMAIL_MESSAGE%(rpt_folder)) # with variables:%(rpt_folder)
            # send_email(emailImage, RECIPIENT_LIST, EMAIL_SUBJECT, EMAIL_MESSAGE) # without variables
            send_email(emailImage,RECIPIENT_LISTS[0],RECIPIENT_LISTS[1],RECIPIENT_LISTS[2], EMAIL_SUBJECT, EMAIL_MESSAGE)

    # Delete all temp folders/files & address book
    delete_temp_folder(REPORT_EMAIL_FOLDER)
    delete_address_book()

